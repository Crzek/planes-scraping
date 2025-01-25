import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill


"""
Aplicar estilos a un archivo de Excel
"""


def apply_styles(celda):
    """
    Aplica boderes y alineación a una celda de Excel

    args:
        celda : openpyxl.cell.cell.Cell
    return : None
    """

    # Aplicar un borde grueso alrededor de la celda
    border_t = "thin"
    border = Border(left=Side(style=border_t),
                    right=Side(style=border_t),
                    top=Side(style=border_t),
                    bottom=Side(style=border_t))
    celda.border = border
    # Aplicar alignment a la celda
    celda.alignment = Alignment(horizontal='center', vertical='center')
    celda.font = Font(name='Arial', size=14)


def apply_styles_font(ws: Workbook = None, col: str = None, row: int = None):
    """
    Aplica negrita en col 1 o fila A

    Args:
        ws (Workbook, optional): _description_. Defaults to None.
        col (str, optional): _description_. Defaults to None.
        row (int, optional): _description_. Defaults to None.
    """
    def apply(col_o_row):
        for celda in ws[col_o_row]:
            celda.font = Font(name="Calibri", bold=True, size=17)

    if col is not None:
        apply(col)
    if row is not None:
        apply(row)


def print_doc(ws, title: str, print_area: str = 'A1:P25'):
    # u ORIENTATION_PORTRAIT para vertical
    # ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    # # ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.oddHeader.center.text = title
    # # Definir el área de impresión como un rango de celdas (por ejemplo, A1:D20)
    # ws.print_area = 'A1:D20'
    # Configurar las opciones de impresión
    ws.print_area = print_area
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    print("Configuración de impresión aplicada con éxito")


def main_styles(today: bool = False):
    from globals import TODAY, TOMORROW, PATH_STATIC_DATA

    # data_url = 'src/app/data/'
    # file = data_url + \
    #     f'vuelos-{TODAY if today else TOMORROW }.xlsx'

    file = f"{PATH_STATIC_DATA}vuelos-{TODAY if today else TOMORROW}.xlsx"
    # abrir un libro de trabajo
    print(f"Aplicando estilos al archivo de excel:\n{file}")
    wb = load_workbook(file)
    # seleccionar la hoja de trabajo
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column

    # Iterar sobre todas las celdas desde A1 hasta la celda máxima
    # con el min_row=3 se salta la fila 1 y 2
    for fila in ws.iter_rows(min_row=3, max_row=max_row, min_col=1, max_col=max_col):
        # Ajustar la altura de la fila 1 a 40 (valor en puntos)
        if fila[0].row == 0:
            continue
        ws.row_dimensions[fila[0].row].height = 20
        for celda in fila:
            apply_styles(celda)

    # hacer mas hacho 1r columna
    ws.column_dimensions['A'].width = 13
    # aplicar estilos a la fila 1
    apply_styles_font(ws, col='A', row=3)

    # aplicar estilos a la fila 1
    # poner nombre en fila A1
    ws['A1'] = TODAY if today else TOMORROW
    # ws.row_dimensions[1].height = 20  # Altura de la fila
    ws["A1"].font = Font(bold=True, size=11)  # Negrita y tamaño

    # definir config impresion
    print_doc(ws, f"Vuelos {TODAY if today else TOMORROW}")

    wb.save(file)
    print("Estilos aplicados con éxito")


if __name__ == "__main__":
    main_styles()
